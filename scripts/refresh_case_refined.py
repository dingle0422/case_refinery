#!/usr/bin/env python3
"""按新 refine 逻辑批量重刷 LanceDB 中已有 refined_knowledge。

三阶段（建议顺序）::

    # 1) 从 LanceDB 导出待重刷记录到 xlsx
    python scripts/refresh_case_refined.py export -o data/case_refined_refresh.xlsx

    # 2) 本地重跑 refine，结果写入 xlsx（refined_knowledge_new 等列）
    python scripts/refresh_case_refined.py refine -i data/case_refined_refresh.xlsx

    # 3) 人工审阅 xlsx 后，回写 LanceDB（仅 merge metadata + refined_knowledge）
    python scripts/refresh_case_refined.py upload -i data/case_refined_refresh.xlsx --confirm

默认筛选（6.2 之前入库的数据）：``ingest_ts`` 严格早于 2026-06-02 00:00
（Asia/Shanghai）、``refine_status=refined``、未 tombstone。

依赖环境变量与主服务相同（``CASE_REFINERY_*``）。xlsx 读写需要 ``openpyxl``。
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Literal
from zoneinfo import ZoneInfo

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_CASE_REFINERY_DIR = os.path.abspath(os.path.join(_SCRIPT_DIR, ".."))
_REPO_ROOT = os.path.abspath(os.path.join(_CASE_REFINERY_DIR, ".."))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from case_refinery.config import Settings, get_settings
from case_refinery.pipeline import classifier, refiner
from case_refinery.pipeline.lancedb_client import (
    LanceDBError,
    LanceDBV2Client,
    _coerce_lance_document_id,
)
from case_refinery.pipeline.runner import _now_ms
from case_refinery.pipeline.upstream import CaseDict

logger = logging.getLogger(__name__)

# openpyxl 不允许写入 XML 非法控制字符（保留 \t \n \r）
_ILLEGAL_XLSX_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

TZ_SHANGHAI = ZoneInfo("Asia/Shanghai")
DEFAULT_CUTOFF = date(2026, 6, 2)

DateOp = Literal["before", "on_or_after"]
DateField = Literal["ingest_ts", "refined_at"]

# xlsx 列顺序（export 写入；refine/upload 在此基础上追加/更新）
COLUMNS: list[str] = [
    "kh_code",
    "document_id",
    "collection_id",
    "questionContent",
    "originalAnswer",
    "originalThinking",
    "answerContent",
    "thinking",
    "case_polarity",
    "record_hash",
    "question_hash",
    "refine_status",
    "refine_attempts",
    "refined_at",
    "ingest_ts",
    "refined_knowledge",
    "refined_knowledge_new",
    "refine_ok",
    "refine_error",
    "refined_at_new",
    "upload_ok",
    "upload_error",
]


def _cutoff_ms(cutoff: date, *, end_of_day: bool = False) -> int:
    """cutoff 日 00:00 或 23:59:59.999（上海时区）转 epoch ms。"""
    if end_of_day:
        dt = datetime(
            cutoff.year,
            cutoff.month,
            cutoff.day,
            23,
            59,
            59,
            999000,
            tzinfo=TZ_SHANGHAI,
        )
    else:
        dt = datetime(cutoff.year, cutoff.month, cutoff.day, 0, 0, 0, tzinfo=TZ_SHANGHAI)
    return int(dt.timestamp() * 1000)


def _parse_cutoff(raw: str) -> date:
    try:
        return date.fromisoformat(raw.strip())
    except ValueError as e:
        raise argparse.ArgumentTypeError(
            f"无效日期 {raw!r}，请使用 YYYY-MM-DD"
        ) from e


def _parse_kh_codes(raw: str | None) -> set[str] | None:
    if not raw:
        return None
    codes = {x.strip() for x in raw.split(",") if x.strip()}
    return codes or None


def _coerce_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _match_date_filter(
    ts_ms: int,
    *,
    cutoff_ms: int,
    date_op: DateOp,
) -> bool:
    if ts_ms <= 0:
        return False
    if date_op == "before":
        return ts_ms < cutoff_ms
    return ts_ms >= cutoff_ms


async def list_case_collections(
    cli: LanceDBV2Client,
    settings: Settings,
    *,
    kh_filter: set[str] | None = None,
) -> list[tuple[str, str]]:
    data = await cli._request("GET", "/v2/collections")
    rows: list[tuple[str, str]] = []
    prefix = settings.lancedb_collection_prefix
    for item in (data or {}).get("collections") or []:
        if not isinstance(item, dict):
            continue
        cid = str(item.get("collection_id") or "")
        if not cid.startswith(prefix):
            continue
        kh_code = cid[len(prefix):]
        if kh_filter is not None and kh_code not in kh_filter:
            continue
        rows.append((kh_code, cid))
    rows.sort(key=lambda x: x[0])
    return rows


@dataclass
class ExportRow:
    kh_code: str
    document_id: str
    collection_id: str
    case: CaseDict
    case_polarity: str
    record_hash: str
    question_hash: str
    refine_status: str
    refine_attempts: int
    refined_at: int
    ingest_ts: int
    refined_knowledge: str

    def as_xlsx_dict(self) -> dict[str, Any]:
        return {
            "kh_code": self.kh_code,
            "document_id": self.document_id,
            "collection_id": self.collection_id,
            "questionContent": self.case.get("questionContent") or "",
            "originalAnswer": self.case.get("originalAnswer") or "",
            "originalThinking": self.case.get("originalThinking") or "",
            "answerContent": self.case.get("answerContent") or "",
            "thinking": self.case.get("thinking") or "",
            "case_polarity": self.case_polarity,
            "record_hash": self.record_hash,
            "question_hash": self.question_hash,
            "refine_status": self.refine_status,
            "refine_attempts": self.refine_attempts,
            "refined_at": self.refined_at,
            "ingest_ts": self.ingest_ts,
            "refined_knowledge": self.refined_knowledge,
            "refined_knowledge_new": "",
            "refine_ok": "",
            "refine_error": "",
            "refined_at_new": "",
            "upload_ok": "",
            "upload_error": "",
        }


async def fetch_documents_for_export(
    cli: LanceDBV2Client,
    kh_code: str,
    collection_id: str,
    *,
    page_size: int,
    cutoff_ms: int,
    date_op: DateOp,
    date_field: DateField,
    only_refined: bool,
) -> list[ExportRow]:
    rows: list[ExportRow] = []
    offset = 0

    while True:
        params = {
            "include_content": "true",
            "limit": page_size,
            "offset": offset,
        }
        data = await cli._request(
            "GET",
            f"/v2/collections/{collection_id}/documents",
            params=params,
        )
        docs = (data or {}).get("documents") or []
        n = len(docs)

        for doc in docs:
            if not isinstance(doc, dict):
                continue
            md = doc.get("metadata") or {}
            if bool(md.get("tombstoned", False)):
                continue

            refine_status = str(md.get("refine_status") or "")
            if only_refined and refine_status != "refined":
                continue

            ts_ms = _coerce_int(md.get(date_field))
            if not _match_date_filter(ts_ms, cutoff_ms=cutoff_ms, date_op=date_op):
                continue

            doc_id = str(doc.get("document_id") or "")
            if not doc_id:
                continue

            case: CaseDict = {
                "questionContent": str(
                    md.get("question_content") or doc.get("content") or ""
                ),
                "originalAnswer": str(md.get("original_answer") or ""),
                "originalThinking": str(md.get("original_thinking") or ""),
                "answerContent": str(md.get("answer_content") or ""),
                "thinking": str(md.get("thinking") or ""),
            }
            polarity = str(md.get("case_polarity") or classifier.classify(case))

            rows.append(
                ExportRow(
                    kh_code=str(md.get("kh_code") or kh_code),
                    document_id=doc_id,
                    collection_id=collection_id,
                    case=case,
                    case_polarity=polarity,
                    record_hash=str(md.get("record_hash") or ""),
                    question_hash=str(md.get("question_hash") or ""),
                    refine_status=refine_status,
                    refine_attempts=_coerce_int(md.get("refine_attempts")),
                    refined_at=_coerce_int(md.get("refined_at")),
                    ingest_ts=_coerce_int(md.get("ingest_ts")),
                    refined_knowledge=str(md.get("refined_knowledge") or ""),
                )
            )

        if n < page_size:
            break
        offset += page_size

    return rows


def _sanitize_xlsx_value(v: Any) -> Any:
    if isinstance(v, str):
        return _ILLEGAL_XLSX_CHAR_RE.sub("", v)
    return v


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise SystemExit(
            "缺少 openpyxl，请执行: pip install openpyxl"
        ) from e


def write_xlsx(path: str, rows: list[dict[str, Any]]) -> None:
    _require_openpyxl()
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "cases"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([_sanitize_xlsx_value(row.get(c, "")) for c in COLUMNS])
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb.save(path)
    logger.info("已写入 %d 行 -> %s", len(rows), os.path.abspath(path))


def read_xlsx(path: str) -> list[dict[str, Any]]:
    _require_openpyxl()
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if not headers:
        wb.close()
        return []
    col_index = {str(h): i for i, h in enumerate(headers) if h}

    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None or v == "" for v in r):
            continue
        item = {}
        for name in COLUMNS:
            idx = col_index.get(name)
            if idx is None:
                item[name] = ""
            else:
                v = r[idx] if idx < len(r) else ""
                item[name] = "" if v is None else v
        rows.append(item)
    wb.close()
    return rows


def _row_to_case(row: dict[str, Any]) -> CaseDict:
    return {
        "questionContent": str(row.get("questionContent") or ""),
        "originalAnswer": str(row.get("originalAnswer") or ""),
        "originalThinking": str(row.get("originalThinking") or ""),
        "answerContent": str(row.get("answerContent") or ""),
        "thinking": str(row.get("thinking") or ""),
    }


async def cmd_export(args: argparse.Namespace, settings: Settings) -> int:
    kh_filter = _parse_kh_codes(args.kh_codes)
    cutoff_ms = _cutoff_ms(args.cutoff_date)
    cli = LanceDBV2Client(settings=settings)
    all_rows: list[dict[str, Any]] = []

    try:
        collections = await list_case_collections(cli, settings, kh_filter=kh_filter)
        if not collections:
            logger.warning("未找到 case_* collection")
            return 1

        for kh_code, collection_id in collections:
            logger.info(
                "[export] %s 拉取中 date_field=%s date_op=%s cutoff=%s",
                collection_id,
                args.date_field,
                args.date_op,
                args.cutoff_date,
            )
            batch = await fetch_documents_for_export(
                cli,
                kh_code,
                collection_id,
                page_size=settings.lancedb_list_page_size,
                cutoff_ms=cutoff_ms,
                date_op=args.date_op,
                date_field=args.date_field,
                only_refined=not args.include_raw_fallback,
            )
            logger.info("[export] %s 命中 %d 条", collection_id, len(batch))
            all_rows.extend(r.as_xlsx_dict() for r in batch)
    except LanceDBError as e:
        logger.error("LanceDB 调用失败: %s", e)
        return 1
    finally:
        await cli.aclose()

    write_xlsx(args.output, all_rows)
    print(
        f"export 完成: {len(all_rows)} 条 -> {os.path.abspath(args.output)}\n"
        f"  筛选: {args.date_field} {args.date_op} {args.cutoff_date} "
        f"(cutoff_ms={cutoff_ms})"
    )
    return 0


async def _refine_one_row(
    row: dict[str, Any],
    *,
    settings: Settings,
    sem: asyncio.Semaphore,
) -> dict[str, Any]:
    async with sem:
        case = _row_to_case(row)
        polarity = str(row.get("case_polarity") or classifier.classify(case))
        if polarity not in ("positive", "negative"):
            polarity = classifier.classify(case)

        result = await refiner.refine(
            case,
            polarity,  # type: ignore[arg-type]
            settings=settings,
        )
        out = dict(row)
        if result.ok:
            out["refined_knowledge_new"] = result.refined_knowledge
            out["refine_ok"] = "true"
            out["refine_error"] = ""
            out["refined_at_new"] = _now_ms()
        else:
            out["refined_knowledge_new"] = ""
            out["refine_ok"] = "false"
            out["refine_error"] = result.error
            out["refined_at_new"] = 0
        return out


async def _refine_one_row_indexed(
    idx: int,
    row: dict[str, Any],
    *,
    settings: Settings,
    sem: asyncio.Semaphore,
) -> tuple[int, dict[str, Any]]:
    updated = await _refine_one_row(row, settings=settings, sem=sem)
    return idx, updated


def _flush_logs() -> None:
    for handler in logging.root.handlers:
        handler.flush()
    sys.stdout.flush()
    sys.stderr.flush()


async def cmd_refine(args: argparse.Namespace, settings: Settings) -> int:
    path = os.path.abspath(args.input)
    if not os.path.isfile(path):
        logger.error("文件不存在: %s", path)
        return 1

    logger.info("[refine] 读取 xlsx: %s", path)
    _flush_logs()
    rows = read_xlsx(path)
    if not rows:
        logger.warning("xlsx 无数据行")
        return 1

    if args.limit > 0:
        rows = rows[: args.limit]

    total = len(rows)
    out_path = os.path.abspath(args.output or path)
    working = [dict(r) for r in rows]
    sem = asyncio.Semaphore(max(1, args.concurrency))
    progress_every = max(1, args.progress_every)
    checkpoint_every = max(0, args.checkpoint_every)

    logger.info(
        "[refine] 开始处理 %d 条，并发=%d，进度间隔=%d，checkpoint=%s",
        total,
        args.concurrency,
        progress_every,
        checkpoint_every if checkpoint_every else "关闭",
    )
    _flush_logs()

    tasks = [
        asyncio.create_task(
            _refine_one_row_indexed(i, row, settings=settings, sem=sem)
        )
        for i, row in enumerate(rows)
    ]
    ok_count = 0
    fail_count = 0
    done = 0

    for coro in asyncio.as_completed(tasks):
        idx, row = await coro
        working[idx] = row
        done += 1
        if str(row.get("refine_ok")).lower() == "true":
            ok_count += 1
        else:
            fail_count += 1

        if done % progress_every == 0 or done == total:
            logger.info(
                "[refine] 进度 %d/%d ok=%d fail=%d",
                done, total, ok_count, fail_count,
            )
            _flush_logs()

        if checkpoint_every and (done % checkpoint_every == 0 or done == total):
            write_xlsx(out_path, working)
            logger.info("[refine] checkpoint 已写入 %s（%d/%d）", out_path, done, total)
            _flush_logs()

    if not checkpoint_every:
        write_xlsx(out_path, working)

    logger.info("[refine] 完成 total=%d ok=%d fail=%d", total, ok_count, fail_count)
    _flush_logs()
    print(
        f"refine 完成: total={total} ok={ok_count} fail={fail_count} -> {out_path}",
        flush=True,
    )
    return 0 if fail_count == 0 or not args.strict else 1


def _build_upload_merge_document(row: dict[str, Any]) -> dict:
    kh_code = str(row.get("kh_code") or "")
    doc_id = row.get("document_id")
    refined_knowledge = str(row.get("refined_knowledge_new") or "")
    refined_at_new = _coerce_int(row.get("refined_at_new"), _now_ms())
    attempts = _coerce_int(row.get("refine_attempts"), 1)

    return {
        "document_id": _coerce_lance_document_id(doc_id),
        "content": "",
        "content_tokenized": "",
        "vector": [],
        "metadata": {
            "kh_code": kh_code,
            "source": "case_refinery",
            "record_hash": str(row.get("record_hash") or ""),
            "question_hash": str(row.get("question_hash") or ""),
            "refined_knowledge": refined_knowledge,
            "refine_status": "refined",
            "refine_attempts": attempts,
            "refined_at": refined_at_new,
            "ingest_ts": _now_ms(),
            "schema_version": 1,
            "tombstoned": False,
        },
    }


async def cmd_upload(args: argparse.Namespace, settings: Settings) -> int:
    if not args.confirm:
        print(
            "upload 需要显式确认：加上 --confirm 才会写入 LanceDB。\n"
            "可先不加 --confirm 查看将上传多少条（dry-run）。"
        )

    path = os.path.abspath(args.input)
    if not os.path.isfile(path):
        logger.error("文件不存在: %s", path)
        return 1

    rows = read_xlsx(path)
    candidates = [
        r
        for r in rows
        if str(r.get("refine_ok")).lower() == "true"
        and str(r.get("refined_knowledge_new") or "").strip()
    ]
    if args.limit > 0:
        candidates = candidates[: args.limit]

    if not candidates:
        logger.warning("没有 refine_ok=true 且 refined_knowledge_new 非空的行可上传")
        return 1

    print(f"待上传 {len(candidates)} 条（dry_run={not args.confirm}）")
    if not args.confirm:
        for r in candidates[:5]:
            print(
                f"  - kh={r.get('kh_code')} doc_id={r.get('document_id')} "
                f"len_new={len(str(r.get('refined_knowledge_new') or ''))}"
            )
        if len(candidates) > 5:
            print(f"  ... 另有 {len(candidates) - 5} 条")
        return 0

    cli = LanceDBV2Client(settings=settings)
    ok = 0
    fail = 0
    try:
        for i, row in enumerate(candidates, 1):
            kh_code = str(row.get("kh_code") or "")
            doc = _build_upload_merge_document(row)
            try:
                await cli.upsert_one(kh_code, doc, mode="merge_by_chunk_id")
                row["upload_ok"] = "true"
                row["upload_error"] = ""
                ok += 1
            except LanceDBError as e:
                row["upload_ok"] = "false"
                row["upload_error"] = str(e)
                fail += 1
                logger.warning(
                    "[upload] 失败 kh=%s doc_id=%s: %s",
                    kh_code,
                    row.get("document_id"),
                    e,
                )
            if i % max(1, args.progress_every) == 0 or i == len(candidates):
                logger.info("[upload] 进度 %d/%d ok=%d fail=%d", i, len(candidates), ok, fail)
            # 节流，避免 LanceDB 突发
            if args.upload_delay_s > 0:
                await asyncio.sleep(args.upload_delay_s)
    finally:
        await cli.aclose()

    # 把 upload 状态写回完整 xlsx
    by_key = {
        (str(r.get("kh_code")), str(r.get("document_id"))): r for r in candidates
    }
    for row in rows:
        key = (str(row.get("kh_code")), str(row.get("document_id")))
        if key in by_key:
            src = by_key[key]
            row["upload_ok"] = src.get("upload_ok", "")
            row["upload_error"] = src.get("upload_error", "")

    write_xlsx(path, rows)
    print(f"upload 完成: ok={ok} fail={fail}，状态已写回 {path}")
    return 0 if fail == 0 else 1


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="导出 / 重刷 refine / 回写 LanceDB refined_knowledge",
    )
    parser.add_argument(
        "--log-level",
        default=os.getenv("CASE_REFINERY_LOG_LEVEL", "INFO"),
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_export = sub.add_parser("export", help="从 LanceDB 导出到 xlsx")
    p_export.add_argument(
        "-o",
        "--output",
        default="data/case_refined_refresh.xlsx",
        help="输出 xlsx 路径",
    )
    p_export.add_argument("--kh-codes", default="", help="逗号分隔，仅处理指定 khCode")
    p_export.add_argument(
        "--cutoff-date",
        type=_parse_cutoff,
        default=DEFAULT_CUTOFF,
        help="日期阈值 YYYY-MM-DD（默认 2026-06-02）",
    )
    p_export.add_argument(
        "--date-field",
        choices=("ingest_ts", "refined_at"),
        default="ingest_ts",
        help="用于筛选的时间字段（默认 ingest_ts，即入库/upsert 时间）",
    )
    p_export.add_argument(
        "--date-op",
        choices=("before", "on_or_after"),
        default="before",
        help="before=早于 cutoff；on_or_after=>= cutoff 日 00:00（上海时区）",
    )
    p_export.add_argument(
        "--include-raw-fallback",
        action="store_true",
        help="同时导出 refine_status=raw_fallback 的记录",
    )

    p_refine = sub.add_parser("refine", help="对 xlsx 内 case 重跑 LLM refine")
    p_refine.add_argument("-i", "--input", required=True, help="输入 xlsx")
    p_refine.add_argument(
        "-o",
        "--output",
        default="",
        help="输出 xlsx（默认覆盖 --input）",
    )
    p_refine.add_argument("--concurrency", type=int, default=3, help="并发 refine 数")
    p_refine.add_argument("--limit", type=int, default=0, help="仅处理前 N 条（0=全部）")
    p_refine.add_argument(
        "--progress-every",
        type=int,
        default=10,
        help="每处理多少条打一条进度日志",
    )
    p_refine.add_argument(
        "--checkpoint-every",
        type=int,
        default=200,
        help="每处理多少条写一次 xlsx checkpoint（0=仅结束时写入）",
    )
    p_refine.add_argument(
        "--strict",
        action="store_true",
        help="若存在 refine 失败则 exit 1",
    )

    p_upload = sub.add_parser("upload", help="将审阅后的 refined_knowledge 写回 LanceDB")
    p_upload.add_argument("-i", "--input", required=True, help="输入 xlsx")
    p_upload.add_argument(
        "--confirm",
        action="store_true",
        help="确认写入 LanceDB（无此参数仅 dry-run）",
    )
    p_upload.add_argument("--limit", type=int, default=0)
    p_upload.add_argument("--progress-every", type=int, default=20)
    p_upload.add_argument(
        "--upload-delay-s",
        type=float,
        default=0.05,
        help="每条 upload 间隔秒数",
    )

    return parser


class _FlushingStreamHandler(logging.StreamHandler):
    """管道重定向（如 tee）时避免日志长时间不落盘。"""

    def emit(self, record: logging.LogRecord) -> None:
        super().emit(record)
        self.flush()


async def async_main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    level = getattr(logging, str(args.log_level).upper(), logging.INFO)
    handler = _FlushingStreamHandler(sys.stderr)
    handler.setFormatter(
        logging.Formatter(
            "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    logging.basicConfig(level=level, handlers=[handler], force=True)
    settings = get_settings()

    if args.command == "export":
        return await cmd_export(args, settings)
    if args.command == "refine":
        return await cmd_refine(args, settings)
    if args.command == "upload":
        return await cmd_upload(args, settings)
    parser.error(f"未知子命令: {args.command}")
    return 2


def main(argv: list[str] | None = None) -> int:
    return asyncio.run(async_main(argv))


if __name__ == "__main__":
    raise SystemExit(main())
